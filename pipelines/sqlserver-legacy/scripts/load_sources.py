from __future__ import annotations

import csv
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Tuple

from openpyxl import load_workbook

from .common import (
    extract_date_token,
    iter_chunked,
    normalize_code,
    normalize_text,
    parse_yyyymmdd,
    resolve_path,
    to_float,
)


@dataclass
class SourceFiles:
    stock_csv: str
    pool_csv: str
    sales_csv: str
    product_info_xlsx: str


def _open_csv(path: str):
    return open(path, "r", encoding="gbk", newline="")


def _read_csv_header(path: str) -> List[str]:
    with _open_csv(path) as f:
        reader = csv.reader(f)
        return [normalize_text(x) for x in next(reader, [])]


def _latest_path(paths: List[Path]) -> str:
    if not paths:
        return ""
    return str(max(paths, key=lambda p: p.stat().st_mtime))


def detect_source_files(config: Dict, logger) -> SourceFiles:
    base_dir = config.get("_base_dir")
    data_dir = resolve_path(config.get("paths", {}).get("data_dir", "."), base_dir)
    src_cfg = config.get("sources", {})

    explicit_stock = resolve_path(src_cfg.get("stock_file", ""), data_dir)
    explicit_pool = resolve_path(src_cfg.get("pool_file", ""), data_dir)
    explicit_sales = resolve_path(src_cfg.get("sales_file", ""), data_dir)
    explicit_product = resolve_path(src_cfg.get("product_info_file", ""), data_dir)

    stock_csv = explicit_stock if explicit_stock and Path(explicit_stock).exists() else ""
    pool_csv = explicit_pool if explicit_pool and Path(explicit_pool).exists() else ""
    sales_csv = explicit_sales if explicit_sales and Path(explicit_sales).exists() else ""
    product_info = explicit_product if explicit_product and Path(explicit_product).exists() else ""

    csv_paths = [p for p in Path(data_dir).glob("*.csv") if p.is_file()]

    stock_candidates: List[Path] = []
    pool_candidates: List[Path] = []
    sales_candidates: List[Path] = []

    for p in csv_paths:
        try:
            header = _read_csv_header(str(p))
        except Exception:
            continue
        text = "|".join(header)
        if ("吊牌金额" in text) and ("店铺名称" in text):
            sales_candidates.append(p)
            continue
        if ("尺码" in text) and ("分配池" in text) and ("可用数" in text):
            stock_candidates.append(p)
            continue
        if ("分配池" in text) and ("可用数" in text):
            pool_candidates.append(p)

    if not stock_csv:
        stock_csv = _latest_path(stock_candidates)
    if not pool_csv:
        pool_csv = _latest_path(pool_candidates)
    if not sales_csv:
        sales_csv = _latest_path(sales_candidates)

    if not product_info:
        xlsx_candidates = [p for p in Path(data_dir).glob("*.xlsx") if p.is_file()]
        product_info = _latest_path(xlsx_candidates)

    if not stock_csv or not pool_csv or not sales_csv or not product_info:
        raise FileNotFoundError(
            f"missing source files: stock={bool(stock_csv)} pool={bool(pool_csv)} sales={bool(sales_csv)} product={bool(product_info)}"
        )

    logger.info(
        "discover",
        "source files detected",
        {
            "stock_csv": Path(stock_csv).name,
            "pool_csv": Path(pool_csv).name,
            "sales_csv": Path(sales_csv).name,
            "product_info_xlsx": Path(product_info).name,
        },
    )

    return SourceFiles(
        stock_csv=stock_csv,
        pool_csv=pool_csv,
        sales_csv=sales_csv,
        product_info_xlsx=product_info,
    )


def detect_report_week(stock_csv_path: str) -> dt.date:
    token = extract_date_token(Path(stock_csv_path).name)
    parsed = parse_yyyymmdd(token)
    if parsed:
        return parsed
    return dt.datetime.fromtimestamp(Path(stock_csv_path).stat().st_mtime).date()


def _insert_chunked(conn, sql: str, rows: Iterable[Tuple], chunk_size: int = 5000) -> int:
    cur = conn.cursor()
    cur.fast_executemany = True
    total = 0
    for chunk in iter_chunked(rows, size=chunk_size):
        cur.executemany(sql, chunk)
        total += len(chunk)
    conn.commit()
    return total


def _iter_stock_rows(path: str, report_week: dt.date) -> Iterator[Tuple]:
    source_file = Path(path).name
    with _open_csv(path) as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if len(row) < 4:
                continue
            pool = normalize_text(row[0])
            sku = normalize_code(row[1])
            size_name = normalize_text(row[2])
            qty = to_float(row[3])
            if not sku:
                continue
            yield (report_week, pool, sku, size_name, qty, source_file)


def _iter_pool_rows(path: str, report_week: dt.date) -> Iterator[Tuple]:
    source_file = Path(path).name
    with _open_csv(path) as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if len(row) < 3:
                continue
            pool = normalize_text(row[0])
            sku = normalize_code(row[1])
            qty = to_float(row[2])
            if not sku:
                continue
            yield (report_week, pool, sku, qty, source_file)


def _iter_sales_rows(path: str, report_week: dt.date) -> Iterator[Tuple]:
    source_file = Path(path).name
    with _open_csv(path) as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if len(row) < 7:
                continue
            settle_date = parse_yyyymmdd(row[0])
            doc_type = normalize_text(row[1])
            store_name = normalize_text(row[2])
            sku = normalize_code(row[3])
            sales_qty = to_float(row[4])
            sales_amt = to_float(row[5])
            tag_amt = to_float(row[6])
            if not sku:
                continue
            yield (report_week, settle_date, doc_type, store_name, sku, sales_qty, sales_amt, tag_amt, source_file)


def _iter_inventory_latest_rows(path: str, inventory_date: dt.date) -> Iterator[Tuple]:
    source_file = Path(path).name
    with _open_csv(path) as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if len(row) < 4:
                continue
            pool = normalize_text(row[0])
            sku = normalize_code(row[1])
            size_name = normalize_text(row[2])
            qty = to_float(row[3])
            if not sku:
                continue
            yield (inventory_date, pool, sku, size_name, qty, source_file)


def _iter_sales_day_rows(path: str, sales_date: dt.date) -> Iterator[Tuple]:
    source_file = Path(path).name
    with _open_csv(path) as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if len(row) < 7:
                continue
            settle_date = parse_yyyymmdd(row[0])
            if settle_date != sales_date:
                continue
            doc_type = normalize_text(row[1])
            store_name = normalize_text(row[2])
            sku = normalize_code(row[3])
            sales_qty = to_float(row[4])
            sales_amt = to_float(row[5])
            tag_amt = to_float(row[6])
            if not sku:
                continue
            yield (sales_date, settle_date, doc_type, store_name, sku, sales_qty, sales_amt, tag_amt, source_file)


def _normalize_header_name(value) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[\s_\-()（）]", "", text)
    return text


def _resolve_idx(header: List[str], candidates: List[str]) -> Optional[int]:
    lookup = {_normalize_header_name(v): idx for idx, v in enumerate(header) if normalize_text(v)}
    for c in candidates:
        key = _normalize_header_name(c)
        if key in lookup:
            return lookup[key]
    return None


def _pick_product_sheet(workbook) -> List:
    preferred = ["商品主数据(ANTA)", "Sheet2", "Sheet1"]
    picked = []
    for name in preferred:
        if name in workbook.sheetnames:
            picked.append(workbook[name])
    if not picked:
        picked = list(workbook.worksheets)
    return picked


def _find_header_row(ws, max_rows: int = 20) -> Tuple[Optional[int], Optional[List[str]]]:
    for r in range(1, max_rows + 1):
        row = [normalize_text(x) for x in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))]
        if any("货号" in cell for cell in row):
            return r, row
    return None, None


def _load_product_info_records(path: str) -> List[Tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: Dict[str, Dict] = {}
        selected_sheet_name = ""
        for ws in _pick_product_sheet(workbook):
            header_row_no, header = _find_header_row(ws)
            if not header_row_no or not header:
                continue

            idx_sku = _resolve_idx(header, ["货号", "商品货号", "SKU", "sku"])
            if idx_sku is None:
                continue

            idx_style = _resolve_idx(header, ["款号", "款式", "款式编码"])
            idx_major = _resolve_idx(header, ["大类"])
            idx_category = _resolve_idx(header, ["中类", "品类"])
            idx_name = _resolve_idx(header, ["品名", "商品名称"])
            idx_price = _resolve_idx(header, ["吊牌价", "零售价"])
            idx_season = _resolve_idx(header, ["产品季", "产品季节", "季节"])
            idx_gender = _resolve_idx(header, ["性别"])
            idx_story = _resolve_idx(header, ["故事包"])
            idx_color = _resolve_idx(header, ["色系", "颜色"])

            for row in ws.iter_rows(min_row=header_row_no + 1, values_only=True):
                sku = normalize_code(row[idx_sku] if idx_sku < len(row) else "")
                if not sku:
                    continue
                rec = result.setdefault(
                    sku,
                    {
                        "sku": sku,
                        "style": "",
                        "major_category": "",
                        "category": "",
                        "product_name": "",
                        "tag_price": 0.0,
                        "season": "",
                        "gender": "",
                        "story_pack": "",
                        "color": "",
                    },
                )

                def fill_text(key: str, idx: Optional[int]):
                    if idx is None or idx >= len(row):
                        return
                    val = normalize_text(row[idx])
                    if val and not rec[key]:
                        rec[key] = val

                fill_text("style", idx_style)
                fill_text("major_category", idx_major)
                fill_text("category", idx_category)
                fill_text("product_name", idx_name)
                fill_text("season", idx_season)
                fill_text("gender", idx_gender)
                fill_text("story_pack", idx_story)
                fill_text("color", idx_color)

                if idx_price is not None and idx_price < len(row) and rec["tag_price"] == 0:
                    rec["tag_price"] = to_float(row[idx_price])

            selected_sheet_name = ws.title
            if result:
                break

        rows = []
        for sku, rec in result.items():
            style = normalize_code(rec.get("style"))
            if not style and "-" in sku:
                style = sku.split("-", 1)[0]
            rows.append(
                (
                    sku,
                    style,
                    normalize_text(rec.get("major_category")),
                    normalize_text(rec.get("category")),
                    normalize_text(rec.get("product_name")),
                    to_float(rec.get("tag_price")),
                    normalize_text(rec.get("season")),
                    normalize_text(rec.get("gender")),
                    normalize_text(rec.get("story_pack")),
                    normalize_text(rec.get("color")),
                    selected_sheet_name,
                )
            )
        return rows
    finally:
        workbook.close()


def stage_sources(conn, files: SourceFiles, report_week: dt.date, logger) -> Dict[str, int]:
    logger.info("stage", "clean existing staging rows", {"report_week": report_week.isoformat()})
    cur = conn.cursor()
    for table in [
        "stg_stock_daily",
        "stg_pool_stock_daily",
        "stg_sales_daily",
        "stg_product_info",
    ]:
        cur.execute(f"DELETE FROM dbo.{table} WHERE report_week = ?", (report_week,))
    conn.commit()

    stock_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_stock_daily(report_week, pool_name, sku, size_name, available_qty, source_file)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        _iter_stock_rows(files.stock_csv, report_week),
    )
    logger.info("stage", "stock staged", {"rows": stock_count, "file": Path(files.stock_csv).name})

    pool_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_pool_stock_daily(report_week, pool_name, sku, available_qty, source_file)
        VALUES (?, ?, ?, ?, ?)
        """,
        _iter_pool_rows(files.pool_csv, report_week),
    )
    logger.info("stage", "pool stock staged", {"rows": pool_count, "file": Path(files.pool_csv).name})

    sales_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_sales_daily(
            report_week, settlement_date, doc_type, store_name, sku, sales_qty, sales_amt, tag_amt, source_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        _iter_sales_rows(files.sales_csv, report_week),
    )
    logger.info("stage", "sales staged", {"rows": sales_count, "file": Path(files.sales_csv).name})

    product_records = _load_product_info_records(files.product_info_xlsx)
    product_rows = [
        (
            report_week,
            rec[0],
            rec[1],
            rec[2],
            rec[3],
            rec[4],
            rec[5],
            rec[6],
            rec[7],
            rec[8],
            rec[9],
            rec[10],
            Path(files.product_info_xlsx).name,
        )
        for rec in product_records
    ]

    product_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_product_info(
            report_week, sku, style, major_category, category, product_name, tag_price,
            season, gender, story_pack, color, source_sheet, source_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        product_rows,
        chunk_size=3000,
    )
    logger.info(
        "stage",
        "product info staged",
        {"rows": product_count, "file": Path(files.product_info_xlsx).name},
    )

    return {
        "stock_rows": stock_count,
        "pool_rows": pool_count,
        "sales_rows": sales_count,
        "product_rows": product_count,
    }


def stage_daily_sources(
    conn,
    files: SourceFiles,
    sales_date: dt.date,
    inventory_date: dt.date,
    logger,
) -> Dict[str, int]:
    logger.info(
        "stage_daily",
        "clean daily staging rows",
        {"sales_date": sales_date.isoformat(), "inventory_date": inventory_date.isoformat()},
    )
    cur = conn.cursor()
    cur.execute("DELETE FROM dbo.stg_sales_day WHERE sales_date = ?", (sales_date,))
    cur.execute("TRUNCATE TABLE dbo.stg_inventory_latest")
    cur.execute("DELETE FROM dbo.stg_pool_stock_daily WHERE report_week = ?", (inventory_date,))
    cur.execute("DELETE FROM dbo.stg_product_info WHERE report_week = ?", (inventory_date,))
    conn.commit()

    inventory_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_inventory_latest(
            inventory_date, pool_name, sku, size_name, available_qty, source_file
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        _iter_inventory_latest_rows(files.stock_csv, inventory_date),
    )
    logger.info(
        "stage_daily",
        "inventory latest staged",
        {"rows": inventory_count, "file": Path(files.stock_csv).name},
    )

    pool_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_pool_stock_daily(report_week, pool_name, sku, available_qty, source_file)
        VALUES (?, ?, ?, ?, ?)
        """,
        _iter_pool_rows(files.pool_csv, inventory_date),
    )
    logger.info(
        "stage_daily",
        "pool stock staged",
        {"rows": pool_count, "file": Path(files.pool_csv).name, "report_week": inventory_date.isoformat()},
    )

    sales_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_sales_day(
            sales_date, settlement_date, doc_type, store_name, sku, sales_qty, sales_amt, tag_amt, source_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        _iter_sales_day_rows(files.sales_csv, sales_date),
    )
    logger.info(
        "stage_daily",
        "sales day staged",
        {"rows": sales_count, "file": Path(files.sales_csv).name, "sales_date": sales_date.isoformat()},
    )

    product_records = _load_product_info_records(files.product_info_xlsx)
    product_rows = [
        (
            inventory_date,
            rec[0],
            rec[1],
            rec[2],
            rec[3],
            rec[4],
            rec[5],
            rec[6],
            rec[7],
            rec[8],
            rec[9],
            rec[10],
            Path(files.product_info_xlsx).name,
        )
        for rec in product_records
    ]
    product_count = _insert_chunked(
        conn,
        """
        INSERT INTO dbo.stg_product_info(
            report_week, sku, style, major_category, category, product_name, tag_price,
            season, gender, story_pack, color, source_sheet, source_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        product_rows,
        chunk_size=3000,
    )
    logger.info(
        "stage_daily",
        "product info staged",
        {"rows": product_count, "file": Path(files.product_info_xlsx).name, "report_week": inventory_date.isoformat()},
    )

    return {
        "inventory_rows": inventory_count,
        "pool_rows": pool_count,
        "sales_rows": sales_count,
        "product_rows": product_count,
    }
