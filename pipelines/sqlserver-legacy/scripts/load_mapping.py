from __future__ import annotations

import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook

from .common import normalize_text, resolve_path, to_float


def _pick_counter(counter: Counter) -> str:
    if not counter:
        return ""
    top = max(counter.values())
    cands = sorted([k for k, v in counter.items() if v == top], key=lambda x: str(x))
    return cands[0] if cands else ""


def load_mapping_dimensions(conn, config: Dict, logger) -> Dict[str, int]:
    wb_path = resolve_path(config.get("mapping", {}).get("workbook_path", ""), config.get("_base_dir"))
    if not wb_path or not Path(wb_path).exists():
        raise FileNotFoundError(f"mapping workbook not found: {wb_path}")

    workbook = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        if "全渠道仓店" not in workbook.sheetnames:
            raise ValueError("sheet[全渠道仓店] missing in mapping workbook")
        if "品类关联仓" not in workbook.sheetnames:
            raise ValueError("sheet[品类关联仓] missing in mapping workbook")

        ws_channel = workbook["全渠道仓店"]
        ws_ratio = workbook["品类关联仓"]

        pool_channel_counter = defaultdict(Counter)
        pool_remark_counter = defaultdict(Counter)
        store_channel_counter = defaultdict(Counter)
        pool_ratio_counter = defaultdict(Counter)

        for row in ws_channel.iter_rows(min_row=2, values_only=True):
            pool = normalize_text(row[0] if len(row) > 0 else "")
            channel = normalize_text(row[2] if len(row) > 2 else "")
            remark = normalize_text(row[4] if len(row) > 4 else "")
            store = normalize_text(row[6] if len(row) > 6 else "")
            store_channel = normalize_text(row[7] if len(row) > 7 else "")

            if pool and channel:
                pool_channel_counter[pool][channel] += 1
            if pool and remark:
                pool_remark_counter[pool][remark] += 1
            if store and store_channel:
                store_channel_counter[store][store_channel] += 1

        for row in ws_ratio.iter_rows(min_row=2, values_only=True):
            pool = normalize_text(row[8] if len(row) > 8 else "")
            ratio = row[9] if len(row) > 9 else None
            if not pool or ratio in (None, ""):
                continue
            ratio_num = to_float(ratio)
            pool_ratio_counter[pool][ratio_num] += 1

        dim_pool_channel: List[Tuple[str, str, str]] = []
        for pool, counter in pool_channel_counter.items():
            channel = _pick_counter(counter)
            remark = _pick_counter(pool_remark_counter.get(pool, Counter()))
            if channel:
                dim_pool_channel.append((pool, channel, remark))

        dim_store_channel: List[Tuple[str, str]] = []
        for store, counter in store_channel_counter.items():
            channel = _pick_counter(counter)
            if channel:
                dim_store_channel.append((store, channel))

        dim_pool_ratio: List[Tuple[str, float]] = []
        for pool, counter in pool_ratio_counter.items():
            if not counter:
                continue
            top = max(counter.values())
            cands = sorted([k for k, v in counter.items() if v == top])
            dim_pool_ratio.append((pool, float(cands[0])))

    finally:
        workbook.close()

    cur = conn.cursor()
    cur.execute("DELETE FROM dbo.dim_pool_channel")
    cur.execute("DELETE FROM dbo.dim_store_channel")
    cur.execute("DELETE FROM dbo.dim_pool_ratio")

    if dim_pool_channel:
        cur.fast_executemany = True
        cur.executemany(
            "INSERT INTO dbo.dim_pool_channel(pool_name, channel, remark) VALUES (?, ?, ?)",
            dim_pool_channel,
        )
    if dim_store_channel:
        cur.fast_executemany = True
        cur.executemany(
            "INSERT INTO dbo.dim_store_channel(store_name, channel) VALUES (?, ?)",
            dim_store_channel,
        )
    if dim_pool_ratio:
        cur.fast_executemany = True
        cur.executemany(
            "INSERT INTO dbo.dim_pool_ratio(pool_name, sync_ratio) VALUES (?, ?)",
            dim_pool_ratio,
        )

    conn.commit()

    stats = {
        "pool_channel_rows": len(dim_pool_channel),
        "store_channel_rows": len(dim_store_channel),
        "pool_ratio_rows": len(dim_pool_ratio),
    }
    logger.info("mapping", "mapping dimensions loaded", stats)
    return stats


def export_mapping_gaps(conn, report_week: dt.date, runtime_dir: str, logger) -> Dict:
    cur = conn.cursor()

    cur.execute(
        """
        SELECT DISTINCT s.store_name
        FROM dbo.stg_sales_daily s
        LEFT JOIN dbo.dim_store_channel d ON s.store_name = d.store_name
        WHERE s.report_week = ?
          AND ISNULL(s.store_name, '') <> ''
          AND d.store_name IS NULL
        ORDER BY s.store_name
        """,
        (report_week,),
    )
    missing_store = [r[0] for r in cur.fetchall() if r[0]]

    cur.execute(
        """
        WITH pools AS (
            SELECT DISTINCT pool_name FROM dbo.stg_stock_daily WHERE report_week = ?
            UNION
            SELECT DISTINCT pool_name FROM dbo.stg_pool_stock_daily WHERE report_week = ?
        )
        SELECT p.pool_name
        FROM pools p
        LEFT JOIN dbo.dim_pool_channel d ON p.pool_name = d.pool_name
        WHERE ISNULL(p.pool_name, '') <> ''
          AND d.pool_name IS NULL
        ORDER BY p.pool_name
        """,
        (report_week, report_week),
    )
    missing_pool_channel = [r[0] for r in cur.fetchall() if r[0]]

    cur.execute(
        """
        SELECT DISTINCT p.pool_name
        FROM dbo.stg_pool_stock_daily p
        LEFT JOIN dbo.dim_pool_ratio r ON p.pool_name = r.pool_name
        WHERE p.report_week = ?
          AND ISNULL(p.pool_name, '') <> ''
          AND r.pool_name IS NULL
        ORDER BY p.pool_name
        """,
        (report_week,),
    )
    missing_pool_ratio = [r[0] for r in cur.fetchall() if r[0]]

    runtime = Path(runtime_dir)
    runtime.mkdir(parents=True, exist_ok=True)
    token = report_week.strftime("%Y%m%d")
    output_path = runtime / f"mapping_gaps_{token}.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "missing_store_channel"
    ws1.append(["店铺名称", "渠道(待填)"])
    for name in missing_store:
        ws1.append([name, ""])

    ws2 = wb.create_sheet("missing_pool_channel")
    ws2.append(["分配池", "渠道(待填)", "备注(待填)"])
    for name in missing_pool_channel:
        ws2.append([name, "", ""])

    ws3 = wb.create_sheet("missing_pool_ratio")
    ws3.append(["分配池", "同步比例(待填)"])
    for name in missing_pool_ratio:
        ws3.append([name, ""])

    wb.save(output_path)

    result = {
        "path": str(output_path),
        "missing_store_channel": len(missing_store),
        "missing_pool_channel": len(missing_pool_channel),
        "missing_pool_ratio": len(missing_pool_ratio),
    }
    logger.info("mapping", "mapping gaps exported", result)
    return result
