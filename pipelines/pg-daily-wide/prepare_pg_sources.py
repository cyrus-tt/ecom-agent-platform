from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parents[1]
INBOX_DIR = PROJECT_ROOT / "data" / "inbox"
PREPARED_DIR = PROJECT_ROOT / "data" / "prepared"
CSV_ENCODINGS = ("utf-8-sig", "gb18030", "gbk", "utf-16")
PREPARE_MANIFEST_PATH = PREPARED_DIR / "prepare_manifest.json"
INVENTORY_PREPARED_PATH = PREPARED_DIR / "inventory_latest.csv"
SALES_PREPARED_PATH = PREPARED_DIR / "sales_history.csv"
HEADER_PREPARED_PATH = PREPARED_DIR / "daily_wide_target_header.csv"
INVENTORY_MAP_PREPARED_PATH = PREPARED_DIR / "inventory_channel_map.csv"
SALES_MAP_PREPARED_PATH = PREPARED_DIR / "sales_channel_map.csv"
PRODUCT_PREPARED_PATH = PREPARED_DIR / "product_master_current.csv"
SALES_HEADER = [
    "结算日期",
    "单据类型",
    "店铺名称",
    "货号",
    "销售数量",
    "销售金额",
    "吊牌金额",
]
INVENTORY_HEADERS = (
    ["分配池", "货号", "尺码", "可用数"],
    ["分配池", "货号", "可用数"],
)


def is_shadow_copy(path: Path) -> bool:
    name = path.name.lower()
    return "__from_" in name


def detect_csv(path: Path) -> tuple[str, list[str]]:
    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                header = next(reader)
            return encoding, header
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Cannot decode CSV: {path}") from last_error


def find_inventory_csv() -> Path:
    for path in sorted(INBOX_DIR.glob("*.csv")):
        if is_shadow_copy(path):
            continue
        _, header = detect_csv(path)
        if header in INVENTORY_HEADERS:
            return path
    raise FileNotFoundError("Inventory CSV was not found.")


def find_sales_csvs() -> list[Path]:
    matched: list[Path] = []
    for path in sorted(INBOX_DIR.glob("*.csv")):
        if is_shadow_copy(path):
            continue
        _, header = detect_csv(path)
        if len(header) >= 7 and header[:7] == SALES_HEADER:
            matched.append(path)
    return matched


def find_sales_sources() -> list[tuple[str, Path, str | None]]:
    csv_sources = [("csv", path, None) for path in find_sales_csvs()]
    if csv_sources:
        return csv_sources
    return []


def open_workbook(path: Path):
    return load_workbook(path, read_only=True, data_only=True)


def find_mapping_workbook() -> Path:
    for path in sorted(INBOX_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = open_workbook(path)
        except Exception:
            continue
        try:
            if {"库存映射", "销售映射", "数据宽表（表头）"} <= set(wb.sheetnames):
                return path
        finally:
            wb.close()
    raise FileNotFoundError("Mapping workbook was not found.")


def find_product_workbook() -> Path:
    for path in sorted(INBOX_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = open_workbook(path)
        except Exception:
            continue
        try:
            if "商品主数据(ANTA)" in wb.sheetnames:
                return path
        finally:
            wb.close()
    raise FileNotFoundError("Product workbook was not found.")


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def iso_date_from_token(raw: str) -> str:
    token = re.search(r"(20\d{6})", raw)
    if not token:
        raise ValueError(f"Cannot parse YYYYMMDD token from: {raw}")
    return datetime.strptime(token.group(1), "%Y%m%d").date().isoformat()


def load_prepare_manifest() -> dict:
    if not PREPARE_MANIFEST_PATH.exists():
        return {}
    try:
        return json.loads(PREPARE_MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def file_state(path: Path) -> dict[str, object]:
    stat = path.stat()
    return {
        "path": str(path),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
    }


def same_file_state(path: Path, state: dict | None) -> bool:
    if not state or not path.exists():
        return False
    stat = path.stat()
    return (
        str(path) == str(state.get("path"))
        and stat.st_size == state.get("size")
        and stat.st_mtime_ns == state.get("mtime_ns")
    )


def sales_source_states(sources: list[tuple[str, Path, str | None]]) -> list[dict[str, object]]:
    return [
        {
            "source_type": source_type,
            "sheet_name": sheet_name,
            **file_state(source),
        }
        for source_type, source, sheet_name in sources
    ]


def same_sales_sources(
    sources: list[tuple[str, Path, str | None]],
    states: list[dict[str, object]] | None,
) -> bool:
    if states is None or len(sources) != len(states):
        return False
    for (source_type, source, sheet_name), state in zip(sources, states):
        if state.get("source_type") != source_type or state.get("sheet_name") != sheet_name:
            return False
        if not same_file_state(source, state):
            return False
    return True


def outputs_exist(paths: Iterable[Path]) -> bool:
    return all(path.exists() for path in paths)


def write_csv(path: Path, header: Iterable[str], rows: Iterable[Iterable[object]]) -> int:
    count = 0
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(list(header))
        for row in rows:
            writer.writerow(list(row))
            count += 1
    return count


def export_inventory_csv(source: Path, encoding: str) -> int:
    snapshot_date = iso_date_from_token(source.name)

    def rows():
        with source.open("r", encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                pool_name = text(row.get("分配池"))
                sku = text(row.get("货号"))
                if not pool_name or not sku:
                    continue
                yield [
                    snapshot_date,
                    pool_name,
                    sku,
                    text(row.get("尺码")),
                    text(row.get("可用数")),
                    source.name,
                ]

    return write_csv(
        INVENTORY_PREPARED_PATH,
        ["snapshot_date", "pool_name", "sku", "size_code", "available_qty", "source_file"],
        rows(),
    )


def export_sales_csv(sources: list[tuple[str, Path, str | None]]) -> int:
    def rows():
        for source_type, source, sheet_name in sources:
            if source_type == "csv":
                encoding, _ = detect_csv(source)
                with source.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    for row in reader:
                        sales_date = text(row.get("结算日期"))
                        store_name = text(row.get("店铺名称"))
                        sku = text(row.get("货号"))
                        if not sales_date or not store_name or not sku:
                            continue
                        yield [
                            iso_date_from_token(sales_date),
                            text(row.get("单据类型")),
                            store_name,
                            sku,
                            text(row.get("销售数量")),
                            text(row.get("销售金额")),
                            text(row.get("吊牌金额")),
                            source.name,
                        ]
                continue

            wb = load_workbook(source, read_only=False, data_only=True)
            try:
                ws = wb[sheet_name or wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    sales_date = text(row[0] if len(row) > 0 else "")
                    store_name = text(row[2] if len(row) > 2 else "")
                    sku = text(row[3] if len(row) > 3 else "")
                    if not sales_date or not store_name or not sku:
                        continue
                    yield [
                        iso_date_from_token(sales_date),
                        text(row[1] if len(row) > 1 else ""),
                        store_name,
                        sku,
                        text(row[4] if len(row) > 4 else ""),
                        text(row[5] if len(row) > 5 else ""),
                        text(row[6] if len(row) > 6 else ""),
                        source.name,
                    ]
            finally:
                wb.close()

    return write_csv(
        SALES_PREPARED_PATH,
        [
            "sales_date",
            "doc_type",
            "store_name",
            "sku",
            "sales_qty",
            "sales_amount",
            "tag_amount",
            "source_file",
        ],
        rows(),
    )


def write_prepare_manifest(
    inventory_csv: Path | None,
    sales_sources: list[tuple[str, Path, str | None]],
    mapping_workbook: Path,
    product_workbook: Path,
    row_counts: dict[str, int],
) -> None:
    prepared_files = [str(SALES_PREPARED_PATH)]
    if INVENTORY_PREPARED_PATH.exists():
        prepared_files.insert(0, str(INVENTORY_PREPARED_PATH))

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "inventory_sources": [str(inventory_csv)] if inventory_csv else [],
        "sales_sources": [str(source) for _source_type, source, _sheet_name in sales_sources],
        "prepared_files": prepared_files,
        "source_state": {
            "inventory": file_state(inventory_csv) if inventory_csv else None,
            "sales": sales_source_states(sales_sources),
            "mapping": file_state(mapping_workbook),
            "product": file_state(product_workbook),
        },
        "row_counts": row_counts,
    }
    PREPARE_MANIFEST_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_mapping_workbook(source: Path) -> tuple[int, int]:
    wb = open_workbook(source)
    try:
        inventory_sheet = wb["库存映射"]
        sales_sheet = wb["销售映射"]
        header_sheet = wb["数据宽表（表头）"]

        def inventory_rows():
            for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
                pool_name = text(row[0])
                inventory_channel = text(row[1])
                if not pool_name:
                    continue
                yield [pool_name, inventory_channel, source.name]

        def sales_rows():
            for row in sales_sheet.iter_rows(min_row=2, values_only=True):
                store_name = text(row[0])
                sales_channel = text(row[1])
                if not store_name:
                    continue
                yield [store_name, sales_channel, source.name]

        header_row = next(header_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        write_csv(HEADER_PREPARED_PATH, [text(value) for value in header_row], [])

        inventory_count = write_csv(
            INVENTORY_MAP_PREPARED_PATH,
            ["pool_name", "inventory_channel", "source_file"],
            inventory_rows(),
        )
        sales_count = write_csv(
            SALES_MAP_PREPARED_PATH,
            ["store_name", "sales_channel", "source_file"],
            sales_rows(),
        )
        return inventory_count, sales_count
    finally:
        wb.close()


def export_product_workbook(source: Path) -> int:
    wb = open_workbook(source)
    try:
        sheet = wb["商品主数据(ANTA)"]
        header = [text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: position for position, name in enumerate(header)}
        required = ["货号", "款号", "大类", "中类", "品名", "吊牌价", "产品季", "性别", "故事包"]
        missing = [name for name in required if name not in index]
        if missing:
            raise KeyError(f"Product workbook is missing columns: {missing}")

        def rows():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sku = text(row[index["货号"]])
                if not sku:
                    continue
                yield [
                    sku,
                    text(row[index["款号"]]),
                    text(row[index["大类"]]),
                    text(row[index["中类"]]),
                    text(row[index["品名"]]),
                    text(row[index["吊牌价"]]),
                    text(row[index["产品季"]]),
                    text(row[index["性别"]]),
                    text(row[index["故事包"]]),
                    source.name,
                ]

        return write_csv(
            PRODUCT_PREPARED_PATH,
            [
                "sku",
                "style",
                "major_category",
                "category",
                "product_name",
                "tag_price",
                "season",
                "gender",
                "story_pack",
                "source_file",
            ],
            rows(),
        )
    finally:
        wb.close()


def main() -> None:
    PREPARED_DIR.mkdir(exist_ok=True)
    manifest = load_prepare_manifest()
    source_state = manifest.get("source_state", {})
    row_counts = manifest.get("row_counts", {})

    try:
        inventory_csv = find_inventory_csv()
    except FileNotFoundError:
        inventory_csv = None
    sales_sources = find_sales_sources()
    mapping_workbook = find_mapping_workbook()
    product_workbook = find_product_workbook()

    inventory_reused = False
    if inventory_csv and same_file_state(inventory_csv, source_state.get("inventory")) and outputs_exist([INVENTORY_PREPARED_PATH]):
        inventory_reused = True
        inventory_rows = int(row_counts.get("inventory_rows", 0))
    elif inventory_csv:
        inventory_encoding, _ = detect_csv(inventory_csv)
        inventory_rows = export_inventory_csv(inventory_csv, inventory_encoding)
    else:
        inventory_rows = int(row_counts.get("inventory_rows", 0))

    sales_reused = same_sales_sources(sales_sources, source_state.get("sales")) and outputs_exist([SALES_PREPARED_PATH])
    if sales_reused:
        sales_rows = int(row_counts.get("sales_rows", 0))
    else:
        sales_rows = export_sales_csv(sales_sources)

    mapping_outputs = [HEADER_PREPARED_PATH, INVENTORY_MAP_PREPARED_PATH, SALES_MAP_PREPARED_PATH]
    mapping_reused = same_file_state(mapping_workbook, source_state.get("mapping")) and outputs_exist(mapping_outputs)
    if mapping_reused:
        inventory_map_rows = int(row_counts.get("inventory_map_rows", 0))
        sales_map_rows = int(row_counts.get("sales_map_rows", 0))
    else:
        inventory_map_rows, sales_map_rows = export_mapping_workbook(mapping_workbook)

    product_reused = same_file_state(product_workbook, source_state.get("product")) and outputs_exist([PRODUCT_PREPARED_PATH])
    if product_reused:
        product_rows = int(row_counts.get("product_rows", 0))
    else:
        product_rows = export_product_workbook(product_workbook)

    write_prepare_manifest(
        inventory_csv,
        sales_sources,
        mapping_workbook,
        product_workbook,
        {
            "inventory_rows": inventory_rows,
            "sales_rows": sales_rows,
            "inventory_map_rows": inventory_map_rows,
            "sales_map_rows": sales_map_rows,
            "product_rows": product_rows,
        },
    )

    print(f"Prepared inventory rows: {inventory_rows}" + (" (reused)" if inventory_reused else ""))
    if not inventory_csv:
        print("Inventory source file: not found, pipeline will keep the existing database snapshot.")
    print(f"Prepared sales rows: {sales_rows}" + (" (reused)" if sales_reused else ""))
    print(f"Prepared inventory mappings: {inventory_map_rows}" + (" (reused)" if mapping_reused else ""))
    print(f"Prepared sales mappings: {sales_map_rows}" + (" (reused)" if mapping_reused else ""))
    print(f"Prepared product rows: {product_rows}" + (" (reused)" if product_reused else ""))
    print(
        "Sales source files: "
        + str(
            [
                source.name if source_type == "csv" else f"{source.name}#{sheet_name}"
                for source_type, source, sheet_name in sales_sources
            ]
        )
    )
    print(f"Prepare manifest: {PREPARE_MANIFEST_PATH}")
    print(f"Output directory: {PREPARED_DIR}")


if __name__ == "__main__":
    main()
